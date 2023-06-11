import openpyxl

# file = "C: \\SeleniumPractice\\testdata.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active  # (or) sheet=workbook "Data"1 #get active sheet from excel
#
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
# workbook.save(file)

# multiple data
file = "C: \\SeleniumPractice\\testdatal.xisx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # get active sheet from excel  #  (or) sheet=workbook ["Data"]
#  (or) sheet=workbook ["Data"]
sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "smith"
sheet.cell(2, 1).value = 567
sheet.cell(2, 2).value = "john"
sheet.cell(3, 1).value = 567
sheet.cell(3, 2).value = "david"
workbook.save(file)
